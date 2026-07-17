from __future__ import annotations

import math
import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DATE_HEADER_ROW = 5
COLLABORATOR_HEADER = "COLABORADOR (A)"
TOTAL_LABEL = "PRODUÇÃO TOTAL DIA"
IDEAL_LABEL = "PRODUÇÃO IDEAL DIA"
PERCENT_LABEL = "% ENTREGA PRODUTIVA"
DEFAULT_FOLDER_SUFFIXES = [
    Path("OneDrive") / "Documents" / "Nova pasta",
    Path("Documents") / "Nova pasta",
]


def normalize_label(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    replacements = {
        "Ç": "C",
        "Ã": "A",
        "Á": "A",
        "Â": "A",
        "À": "A",
        "É": "E",
        "Ê": "E",
        "Í": "I",
        "Ó": "O",
        "Ô": "O",
        "Õ": "O",
        "Ú": "U",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"\s+", " ", text)


def normalize_formula(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    formula = re.sub(r"\s+", "", value).upper().replace("$", "")
    while formula.startswith("=(") and formula.endswith(")"):
        formula = "=" + formula[2:-1]
    return formula


def parse_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped == "-":
            return None
        normalized = stripped.replace(".", "").replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


class ProdutividadeService:
    @classmethod
    def resolve_source_dir(cls, source_dir: str | None = None) -> Path:
        candidate_env = os.getenv("VENNER_PRODUTIVIDADE_DIR")
        candidates: list[Path] = []

        if source_dir:
            candidates.append(Path(source_dir).expanduser())
        elif candidate_env:
            candidates.append(Path(candidate_env).expanduser())
        else:
            home = Path.home()
            candidates.extend(home / suffix for suffix in DEFAULT_FOLDER_SUFFIXES)

        for candidate in candidates:
            if candidate.exists() and candidate.is_dir():
                return candidate.resolve()

        if source_dir:
            raise FileNotFoundError(f"Pasta de produtividade não encontrada: {source_dir}")

        searched = ", ".join(str(path) for path in candidates)
        raise FileNotFoundError(
            "Nenhuma pasta padrão de produtividade foi encontrada. "
            f"Verifique uma destas localizações: {searched}"
        )

    @classmethod
    def load_dashboard(cls, source_dir: str | None = None) -> dict[str, Any]:
        base_dir = cls.resolve_source_dir(source_dir)
        files = cls._source_files(base_dir)
        if not files:
            raise FileNotFoundError(f"Nenhuma planilha .xlsx encontrada em {base_dir}")

        summaries: list[dict[str, Any]] = []
        daily_rows: list[dict[str, Any]] = []
        collaborator_rows: list[dict[str, Any]] = []
        issues_rows: list[dict[str, Any]] = []

        for path in files:
            workbook = load_workbook(path, read_only=False, data_only=False)
            for ws in workbook.worksheets:
                if "ANALISE" not in normalize_label(ws.title):
                    continue
                summary, daily, collaborators, issues = cls._analyze_sheet(path, ws)
                summaries.append(summary)
                daily_rows.extend(daily)
                collaborator_rows.extend(collaborators)
                issues_rows.extend(issues)

        if not summaries:
            raise ValueError("Nenhuma aba de análise de produtividade foi encontrada nas planilhas.")

        summaries.sort(key=lambda item: item["percentual"], reverse=True)
        collaborator_rows.sort(key=lambda item: item["producao_total"], reverse=True)
        daily_rows.sort(key=lambda item: (item["setor"], item["data"]))

        period = cls._resolve_period(files)
        latest_mtime = max(path.stat().st_mtime for path in files)
        total_real = sum(item["producao_real"] for item in summaries)
        total_theoretical = sum(item["producao_teorica"] for item in summaries)
        total_days = max((item["dias_planejados"] for item in summaries), default=0)

        top_days = sorted(daily_rows, key=lambda item: item["percentual"], reverse=True)[:5]
        bottom_days = sorted(daily_rows, key=lambda item: item["percentual"])[:5]

        return {
            "periodo": period,
            "source_dir": str(base_dir),
            "generated_at": datetime.now().isoformat(),
            "updated_at": datetime.fromtimestamp(latest_mtime).isoformat(),
            "files": [
                {
                    "name": path.name,
                    "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
                }
                for path in files
            ],
            "overview": {
                "setores": len(summaries),
                "dias_planejados": total_days,
                "producao_real_total": total_real,
                "producao_teorica_total": total_theoretical,
                "percentual_total": (total_real / total_theoretical) if total_theoretical else 0.0,
                "issues_total": len(issues_rows),
            },
            "setores": summaries,
            "diario": daily_rows,
            "top_days": top_days,
            "bottom_days": bottom_days,
            "colaboradores": collaborator_rows,
            "issues": issues_rows,
        }

    @classmethod
    def _source_files(cls, base_dir: Path) -> list[Path]:
        files = []
        for path in sorted(base_dir.glob("*.xlsx")):
            normalized = normalize_label(path.stem)
            if normalized.startswith("RESUMO PRODUTIVIDADE"):
                continue
            files.append(path)
        return files

    @classmethod
    def _resolve_period(cls, files: list[Path]) -> str:
        month_pattern = re.compile(
            r"(JANEIRO|FEVEREIRO|MARCO|ABRIL|MAIO|JUNHO|JULHO|AGOSTO|SETEMBRO|OUTUBRO|NOVEMBRO|DEZEMBRO)\s+(\d{4})"
        )
        for path in files:
            match = month_pattern.search(normalize_label(path.stem))
            if match:
                return f"{match.group(1)} {match.group(2)}"
        return "PERÍODO NÃO IDENTIFICADO"

    @classmethod
    def _analyze_sheet(cls, path: Path, ws):
        labels = cls._row_label_map(ws)
        dates = cls._date_columns(ws)
        collaborator_header_row = labels[normalize_label(COLLABORATOR_HEADER)]
        total_row = labels[normalize_label(TOTAL_LABEL)]
        ideal_row = labels[normalize_label(IDEAL_LABEL)]
        percent_row = labels[normalize_label(PERCENT_LABEL)]
        collaborator_start = collaborator_header_row + 1
        collaborator_end = total_row - 1
        sector = str(ws["A4"].value or path.stem).strip()

        issues = cls._detect_formula_issues(
            ws=ws,
            workbook_name=path.name,
            sheet_name=ws.title,
            collaborator_start=collaborator_start,
            collaborator_end=collaborator_end,
            dates=dates,
            total_row=total_row,
            ideal_row=ideal_row,
            percent_row=percent_row,
        )

        collaborator_rows = []
        for row_idx in range(collaborator_start, collaborator_end + 1):
            collaborator_name = ws.cell(row_idx, 1).value
            if collaborator_name in (None, ""):
                continue
            collaborator_rows.append((row_idx, str(collaborator_name).strip()))

        daily_rows: list[dict[str, Any]] = []
        collaborator_out: list[dict[str, Any]] = []
        summary_status_counter: Counter[str] = Counter()
        collaborator_metrics = {
            row_idx: {
                "setor": sector,
                "colaborador": name,
                "producao_total": 0.0,
                "dias_produtivos": 0,
                "dias_zero": 0,
                "dias_sem_registro": 0,
                "dias_status": 0,
                "status_counter": Counter(),
            }
            for row_idx, name in collaborator_rows
        }

        total_actual = 0.0
        total_ideal = 0.0
        days_meeting_target = 0
        days_zero_total = 0

        for col_idx, date_value in dates:
            actual = 0.0
            ideal = parse_number(ws.cell(ideal_row, col_idx).value) or 0.0
            status_notes: list[str] = []

            for row_idx, collaborator_name in collaborator_rows:
                value = ws.cell(row_idx, col_idx).value
                number = parse_number(value)
                metrics = collaborator_metrics[row_idx]

                if number is not None:
                    actual += number
                    metrics["producao_total"] += number
                    if number > 0:
                        metrics["dias_produtivos"] += 1
                    else:
                        metrics["dias_zero"] += 1
                    continue

                label = normalize_label(value)
                if label in ("", "-"):
                    metrics["dias_sem_registro"] += 1
                else:
                    text_value = str(value).strip()
                    metrics["dias_status"] += 1
                    metrics["status_counter"][text_value] += 1
                    status_notes.append(f"{collaborator_name}: {text_value}")
                    summary_status_counter[text_value] += 1

            percentual = (actual / ideal) if ideal else 0.0
            gap = actual - ideal
            total_actual += actual
            total_ideal += ideal

            if actual >= ideal and ideal > 0:
                days_meeting_target += 1
            if actual == 0:
                days_zero_total += 1

            daily_rows.append(
                {
                    "setor": sector,
                    "data": date_value.date().isoformat(),
                    "producao_real": actual,
                    "producao_teorica": ideal,
                    "percentual": percentual,
                    "gap": gap,
                    "ocorrencias": " | ".join(status_notes),
                }
            )

        for metrics in collaborator_metrics.values():
            productive_days = metrics["dias_produtivos"]
            total = metrics["producao_total"]
            collaborator_out.append(
                {
                    "setor": metrics["setor"],
                    "colaborador": metrics["colaborador"],
                    "producao_total": total,
                    "dias_produtivos": productive_days,
                    "media_dia_produtivo": (total / productive_days) if productive_days else 0.0,
                    "dias_zero": metrics["dias_zero"],
                    "dias_sem_registro": metrics["dias_sem_registro"],
                    "dias_status": metrics["dias_status"],
                    "ocorrencias": ", ".join(
                        f"{status} ({count})" for status, count in sorted(metrics["status_counter"].items())
                    ),
                }
            )

        summary = {
            "setor": sector,
            "arquivo": path.name,
            "dias_planejados": len(dates),
            "producao_real": total_actual,
            "producao_teorica": total_ideal,
            "percentual": (total_actual / total_ideal) if total_ideal else 0.0,
            "media_real_dia": (total_actual / len(dates)) if dates else 0.0,
            "meta_media_dia": (total_ideal / len(dates)) if dates else 0.0,
            "dias_batendo_meta": days_meeting_target,
            "dias_sem_producao": days_zero_total,
            "ocorrencias_relevantes": ", ".join(
                f"{status} ({count})" for status, count in sorted(summary_status_counter.items())
            ),
            "issues_count": len(issues),
        }

        return summary, daily_rows, collaborator_out, issues

    @classmethod
    def _row_label_map(cls, ws) -> dict[str, int]:
        mapping: dict[str, int] = {}
        for row_idx in range(1, ws.max_row + 1):
            label = normalize_label(ws.cell(row_idx, 1).value)
            if label:
                mapping[label] = row_idx
        return mapping

    @classmethod
    def _date_columns(cls, ws) -> list[tuple[int, datetime]]:
        columns: list[tuple[int, datetime]] = []
        col = 2
        while True:
            value = ws.cell(DATE_HEADER_ROW, col).value
            if value in (None, ""):
                break
            if not isinstance(value, datetime):
                raise ValueError(
                    f"Cabeçalho de data inesperado em {ws.title}!{get_column_letter(col)}{DATE_HEADER_ROW}: {value!r}"
                )
            columns.append((col, value))
            col += 1
        return columns

    @classmethod
    def _detect_formula_issues(
        cls,
        ws,
        workbook_name: str,
        sheet_name: str,
        collaborator_start: int,
        collaborator_end: int,
        dates: list[tuple[int, datetime]],
        total_row: int,
        ideal_row: int,
        percent_row: int,
    ) -> list[dict[str, Any]]:
        issues: list[dict[str, Any]] = []
        wrong_daily_total: list[tuple[str, str, str]] = []

        for col_idx, _ in dates:
            col_letter = get_column_letter(col_idx)
            expected = f"=SUM({col_letter}{collaborator_start}:{col_letter}{collaborator_end})"
            found = ws.cell(total_row, col_idx).value
            if normalize_formula(found) != normalize_formula(expected):
                wrong_daily_total.append((col_letter, str(found), expected))

        if wrong_daily_total:
            first_letter = wrong_daily_total[0][0]
            last_letter = wrong_daily_total[-1][0]
            cell_ref = f"{first_letter}{total_row}:{last_letter}{total_row}"
            issues.append(
                {
                    "setor": str(ws["A4"].value or workbook_name).strip(),
                    "arquivo": workbook_name,
                    "aba": sheet_name,
                    "celula": cell_ref,
                    "problema": (
                        f"As fórmulas de produção total do dia em {len(wrong_daily_total)} coluna(s) "
                        "não somam todas as linhas de colaboradores."
                    ),
                    "encontrado": f"Exemplo: {first_letter}{total_row} {wrong_daily_total[0][1]}",
                    "esperado": f"Exemplo: {first_letter}{total_row} {wrong_daily_total[0][2]}",
                }
            )

        daily_percent_formulas = [
            normalize_formula(ws.cell(percent_row, col_idx).value)
            for col_idx, _ in dates
        ]
        if any("-100%" in formula for formula in daily_percent_formulas):
            first_date_col = get_column_letter(dates[0][0])
            last_date_col = get_column_letter(dates[-1][0])
            issues.append(
                {
                    "setor": str(ws["A4"].value or workbook_name).strip(),
                    "arquivo": workbook_name,
                    "aba": sheet_name,
                    "celula": f"{first_date_col}{percent_row}:{last_date_col}{percent_row}",
                    "problema": (
                        "Os percentuais diários estão calculando desvio contra a meta "
                        "(subtraindo 100%), não o percentual entregue."
                    ),
                    "encontrado": "Fórmulas do tipo =(real/teórico)-100%",
                    "esperado": "Fórmulas do tipo =real/teórico",
                }
            )

        first_date_col = dates[0][0]
        last_date_col = dates[-1][0]
        monthly_value_col = last_date_col + 2
        expected_monthly_ideal = (
            f"=SUM({get_column_letter(first_date_col)}{ideal_row}:{get_column_letter(last_date_col)}{ideal_row})"
        )
        monthly_ideal_formula = ws.cell(ideal_row, monthly_value_col).value
        if normalize_formula(monthly_ideal_formula) != normalize_formula(expected_monthly_ideal):
            issues.append(
                {
                    "setor": str(ws["A4"].value or workbook_name).strip(),
                    "arquivo": workbook_name,
                    "aba": sheet_name,
                    "celula": f"{get_column_letter(monthly_value_col)}{ideal_row}",
                    "problema": "O total mensal de produção teórica não soma todos os dias da linha de meta.",
                    "encontrado": str(monthly_ideal_formula),
                    "esperado": expected_monthly_ideal,
                }
            )

        return issues

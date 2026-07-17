from __future__ import annotations

import argparse
import math
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DATE_HEADER_ROW = 5
COLLABORATOR_HEADER = "COLABORADOR (A)"
TOTAL_LABEL = "PRODUÇÃO TOTAL DIA"
IDEAL_LABEL = "PRODUÇÃO IDEAL DIA"
PERCENT_LABEL = "% ENTREGA PRODUTIVA"


@dataclass
class FormulaIssue:
    arquivo: str
    aba: str
    celula: str
    problema: str
    encontrado: str
    esperado: str


def normalize_label(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = (
        text.replace("Ç", "C")
        .replace("Ã", "A")
        .replace("Á", "A")
        .replace("Â", "A")
        .replace("É", "E")
        .replace("Ê", "E")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("Ô", "O")
        .replace("Õ", "O")
        .replace("Ú", "U")
    )
    return re.sub(r"\s+", " ", text)


def normalize_formula(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    formula = re.sub(r"\s+", "", value).upper().replace("$", "")
    while formula.startswith("=(") and formula.endswith(")"):
        formula = "=" + formula[2:-1]
    return formula


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped == "-":
            return None
        normalized = stripped.replace(".", "").replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def format_percent(value: float) -> str:
    return f"{value * 100:.1f}%"


def format_number(value: float) -> str:
    return f"{value:,.0f}".replace(",", ".")


def excel_text(value: Any) -> Any:
    if isinstance(value, str) and value.startswith("="):
        return "'" + value
    return value


def row_label_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for row_idx in range(1, ws.max_row + 1):
        label = normalize_label(ws.cell(row_idx, 1).value)
        if label:
            mapping[label] = row_idx
    return mapping


def date_columns(ws) -> list[tuple[int, datetime]]:
    columns: list[tuple[int, datetime]] = []
    col = 2
    while True:
        value = ws.cell(DATE_HEADER_ROW, col).value
        if value in (None, ""):
            break
        if isinstance(value, datetime):
            columns.append((col, value))
        else:
            raise ValueError(
                f"Cabeçalho de data inesperado em {ws.title}!{get_column_letter(col)}{DATE_HEADER_ROW}: {value!r}"
            )
        col += 1
    return columns


def detect_formula_issues(
    ws,
    workbook_name: str,
    sheet_name: str,
    collaborator_start: int,
    collaborator_end: int,
    dates: list[tuple[int, datetime]],
    total_row: int,
    ideal_row: int,
    percent_row: int,
) -> list[FormulaIssue]:
    issues: list[FormulaIssue] = []
    first_date_col = dates[0][0]
    last_date_col = dates[-1][0]
    monthly_label_col = last_date_col + 1
    monthly_value_col = last_date_col + 2

    wrong_daily_total: list[tuple[str, str, str]] = []
    for col_idx, _ in dates:
        col_letter = get_column_letter(col_idx)
        expected = f"=SUM({col_letter}{collaborator_start}:{col_letter}{collaborator_end})"
        found = ws.cell(total_row, col_idx).value
        if normalize_formula(found) != normalize_formula(expected):
            wrong_daily_total.append((col_letter, str(found), expected))

    if wrong_daily_total:
        first_letter = wrong_daily_total[0][0]
        last_letter = wrong_daily_total[-1][0]
        cell_ref = (
            f"{first_letter}{total_row}"
            if len(wrong_daily_total) == 1
            else f"{first_letter}{total_row}:{last_letter}{total_row}"
        )
        first_found = wrong_daily_total[0][1]
        first_expected = wrong_daily_total[0][2]
        issues.append(
            FormulaIssue(
                arquivo=workbook_name,
                aba=sheet_name,
                celula=cell_ref,
                problema=(
                    f"As fórmulas de produção total do dia em {len(wrong_daily_total)} coluna(s) "
                    "não somam todas as linhas de colaboradores."
                ),
                encontrado=f"Exemplo: {first_letter}{total_row} {first_found}",
                esperado=f"Exemplo: {first_letter}{total_row} {first_expected}",
            )
        )

    expected_monthly_total = f"=SUM({get_column_letter(first_date_col)}{total_row}:{get_column_letter(last_date_col)}{total_row})"
    monthly_total_formula = ws.cell(total_row, monthly_value_col).value
    if normalize_formula(monthly_total_formula) != normalize_formula(expected_monthly_total):
        issues.append(
            FormulaIssue(
                arquivo=workbook_name,
                aba=sheet_name,
                celula=f"{get_column_letter(monthly_value_col)}{total_row}",
                problema="O total mensal de produção real não soma todos os dias da linha de total.",
                encontrado=str(monthly_total_formula),
                esperado=expected_monthly_total,
            )
        )

    expected_monthly_ideal = f"=SUM({get_column_letter(first_date_col)}{ideal_row}:{get_column_letter(last_date_col)}{ideal_row})"
    monthly_ideal_formula = ws.cell(ideal_row, monthly_value_col).value
    if normalize_formula(monthly_ideal_formula) != normalize_formula(expected_monthly_ideal):
        issues.append(
            FormulaIssue(
                arquivo=workbook_name,
                aba=sheet_name,
                celula=f"{get_column_letter(monthly_value_col)}{ideal_row}",
                problema="O total mensal de produção teórica não soma todos os dias da linha de meta.",
                encontrado=str(monthly_ideal_formula),
                esperado=expected_monthly_ideal,
            )
        )

    expected_monthly_percent = f"={get_column_letter(monthly_value_col)}{total_row}/{get_column_letter(monthly_value_col)}{ideal_row}"
    monthly_percent_formula = ws.cell(percent_row, monthly_value_col).value
    if normalize_formula(monthly_percent_formula) != normalize_formula(expected_monthly_percent):
        issues.append(
            FormulaIssue(
                arquivo=workbook_name,
                aba=sheet_name,
                celula=f"{get_column_letter(monthly_value_col)}{percent_row}",
                problema="O percentual mensal não está calculado como real dividido pelo teórico.",
                encontrado=str(monthly_percent_formula),
                esperado=expected_monthly_percent,
            )
        )

    daily_percent_formulas = [
        normalize_formula(ws.cell(percent_row, col_idx).value) for col_idx, _ in dates
    ]
    if any("-100%" in formula for formula in daily_percent_formulas):
        issues.append(
            FormulaIssue(
                arquivo=workbook_name,
                aba=sheet_name,
                celula=f"{get_column_letter(first_date_col)}{percent_row}:{get_column_letter(last_date_col)}{percent_row}",
                problema="Os percentuais diários estão calculando desvio contra a meta (subtraindo 100%), não o percentual entregue.",
                encontrado="Fórmulas do tipo =(real/teórico)-100%",
                esperado="Fórmulas do tipo =real/teórico",
            )
        )

    monthly_label = ws.cell(total_row, monthly_label_col).value
    if normalize_label(monthly_label) != normalize_label("Total Produção real"):
        issues.append(
            FormulaIssue(
                arquivo=workbook_name,
                aba=sheet_name,
                celula=f"{get_column_letter(monthly_label_col)}{total_row}",
                problema="A coluna de fechamento mensal esperado não foi encontrada na posição usual.",
                encontrado=str(monthly_label),
                esperado="Total Produção real",
            )
        )

    return issues


def analyze_analysis_sheet(path: Path, ws):
    labels = row_label_map(ws)
    dates = date_columns(ws)
    collaborator_header_row = labels[normalize_label(COLLABORATOR_HEADER)]
    total_row = labels[normalize_label(TOTAL_LABEL)]
    ideal_row = labels[normalize_label(IDEAL_LABEL)]
    percent_row = labels[normalize_label(PERCENT_LABEL)]
    collaborator_start = collaborator_header_row + 1
    collaborator_end = total_row - 1
    sector = str(ws["A4"].value).strip()

    issues = detect_formula_issues(
        ws=ws,
        workbook_name=path.name,
        sheet_name=ws.title,
        collaborator_start=collaborator_start,
        collaborator_end=collaborator_end,
        dates=dates,
        total_row=total_row,
        ideal_row=ideal_row,
        percent_row=percent_row,
    )

    collaborator_rows = []
    for row_idx in range(collaborator_start, collaborator_end + 1):
        collaborator_name = ws.cell(row_idx, 1).value
        if collaborator_name in (None, ""):
            continue
        collaborator_rows.append((row_idx, str(collaborator_name).strip()))

    daily_rows: list[dict[str, Any]] = []
    collaborator_rows_out: list[dict[str, Any]] = []
    summary_status_counter: Counter[str] = Counter()

    collaborator_metrics = {
        row_idx: {
            "setor": sector,
            "colaborador": name,
            "producao_total": 0.0,
            "dias_produtivos": 0,
            "dias_zero": 0,
            "dias_sem_registro": 0,
            "dias_status": 0,
            "status_counter": Counter(),
        }
        for row_idx, name in collaborator_rows
    }

    total_actual = 0.0
    total_ideal = 0.0
    days_meeting_target = 0
    days_zero_total = 0

    for col_idx, date_value in dates:
        actual = 0.0
        ideal = parse_number(ws.cell(ideal_row, col_idx).value) or 0.0
        status_notes: list[str] = []

        for row_idx, collaborator_name in collaborator_rows:
            value = ws.cell(row_idx, col_idx).value
            number = parse_number(value)
            metrics = collaborator_metrics[row_idx]

            if number is not None:
                actual += number
                metrics["producao_total"] += number
                if number > 0:
                    metrics["dias_produtivos"] += 1
                else:
                    metrics["dias_zero"] += 1
                continue

            label = normalize_label(value)
            if label == "":
                metrics["dias_sem_registro"] += 1
            elif label == "-":
                metrics["dias_sem_registro"] += 1
            else:
                metrics["dias_status"] += 1
                metrics["status_counter"][str(value).strip()] += 1
                status_notes.append(f"{collaborator_name}: {str(value).strip()}")
                summary_status_counter[str(value).strip()] += 1

        pct = (actual / ideal) if ideal else 0.0
        gap = actual - ideal
        total_actual += actual
        total_ideal += ideal
        if actual >= ideal and ideal > 0:
            days_meeting_target += 1
        if actual == 0:
            days_zero_total += 1

        daily_rows.append(
            {
                "Setor": sector,
                "Data": date_value.date().isoformat(),
                "Produção real": actual,
                "Produção teórica": ideal,
                "% real x teórica": pct,
                "Gap": gap,
                "Ocorrências": " | ".join(status_notes),
            }
        )

    for metrics in collaborator_metrics.values():
        productive_days = metrics["dias_produtivos"]
        total = metrics["producao_total"]
        collaborator_rows_out.append(
            {
                "Setor": metrics["setor"],
                "Colaborador": metrics["colaborador"],
                "Produção total": total,
                "Dias produtivos": productive_days,
                "Média por dia produtivo": (total / productive_days) if productive_days else 0.0,
                "Dias com zero": metrics["dias_zero"],
                "Dias sem registro": metrics["dias_sem_registro"],
                "Dias com status": metrics["dias_status"],
                "Ocorrências": ", ".join(
                    f"{status} ({count})" for status, count in sorted(metrics["status_counter"].items())
                ),
            }
        )

    summary = {
        "Setor": sector,
        "Arquivo": path.name,
        "Dias úteis planejados": len(dates),
        "Produção real": total_actual,
        "Produção teórica": total_ideal,
        "% real x teórica": (total_actual / total_ideal) if total_ideal else 0.0,
        "Média real por dia": (total_actual / len(dates)) if dates else 0.0,
        "Meta média por dia": (total_ideal / len(dates)) if dates else 0.0,
        "Dias batendo meta": days_meeting_target,
        "Dias sem produção": days_zero_total,
        "Ocorrências relevantes": ", ".join(
            f"{status} ({count})" for status, count in sorted(summary_status_counter.items())
        ),
        "Inconsistências de fórmula": len(issues),
    }

    return summary, daily_rows, collaborator_rows_out, issues


def autosize_columns(ws):
    for column in ws.columns:
        letter = column[0].column_letter
        max_length = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 40)


def style_workbook(output_path: Path):
    wb = load_workbook(output_path)
    percent_columns = {"% real x teórica"}
    number_columns = {
        "Produção real",
        "Produção teórica",
        "Gap",
        "Produção total",
        "Média por dia produtivo",
        "Média real por dia",
        "Meta média por dia",
    }

    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        for cell in ws[1]:
            cell.font = Font(bold=True)
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2):
            for cell, header in zip(row, headers):
                if header in percent_columns and isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0%"
                elif header in number_columns and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        autosize_columns(ws)
        ws.freeze_panes = "A2"

    wb.save(output_path)


def write_report(
    report_path: Path,
    summaries_df: pd.DataFrame,
    issues_df: pd.DataFrame,
):
    top = summaries_df.sort_values("% real x teórica", ascending=False)
    lines = [
        "# Produtividade Real x Teórica - Fevereiro 2026",
        "",
        "## Consolidado",
        "",
        "| Setor | Real | Teórica | % real x teórica | Dias batendo meta | Dias sem produção |",
        "| --- | ---: | ---: | ---: | ---: | ---: |",
    ]

    for _, row in top.iterrows():
        lines.append(
            f"| {row['Setor']} | {format_number(row['Produção real'])} | {format_number(row['Produção teórica'])} | "
            f"{format_percent(row['% real x teórica'])} | {int(row['Dias batendo meta'])} | {int(row['Dias sem produção'])} |"
        )

    lines.extend(
        [
            "",
            "## Pontos de atenção",
            "",
        ]
    )

    if not issues_df.empty:
        for _, row in issues_df.iterrows():
            lines.append(
                f"- {row['Arquivo']} / {row['Aba']} / {row['Celula']}: {row['Problema']} "
                f"(encontrado: `{row['Encontrado']}` | esperado: `{row['Esperado']}`)"
            )
    else:
        lines.append("- Nenhuma inconsistência de fórmula detectada nas abas analisadas.")

    report_path.write_text("\n".join(lines), encoding="utf-8")


def build_output_name(input_dir: Path) -> tuple[Path, Path]:
    spreadsheet_names = sorted(path.stem for path in input_dir.glob("*.xlsx"))
    period = "CONSOLIDADO"
    for name in spreadsheet_names:
        match = re.search(r"(JANEIRO|FEVEREIRO|MARCO|ABRIL|MAIO|JUNHO|JULHO|AGOSTO|SETEMBRO|OUTUBRO|NOVEMBRO|DEZEMBRO)\s+(\d{4})", normalize_label(name))
        if match:
            period = f"{match.group(1)} {match.group(2)}"
            break

    excel_name = f"RESUMO PRODUTIVIDADE {period}.xlsx"
    report_name = f"relatorio_produtividade_{period.lower().replace(' ', '_')}.md"
    return input_dir / excel_name, Path.cwd() / "scratch" / report_name


def main():
    parser = argparse.ArgumentParser(description="Consolida planilhas de produtividade real x teórica.")
    parser.add_argument("input_dir", type=Path, help="Pasta com as planilhas .xlsx")
    parser.add_argument("--output", type=Path, help="Caminho do arquivo Excel de saída")
    parser.add_argument("--report", type=Path, help="Caminho do relatório Markdown de saída")
    args = parser.parse_args()

    input_dir = args.input_dir.resolve()
    if not input_dir.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {input_dir}")

    output_path, report_path = build_output_name(input_dir)
    if args.output:
        output_path = args.output.resolve()
    if args.report:
        report_path = args.report.resolve()

    summaries: list[dict[str, Any]] = []
    daily_rows: list[dict[str, Any]] = []
    collaborator_rows: list[dict[str, Any]] = []
    issues_rows: list[dict[str, Any]] = []

    for path in sorted(input_dir.glob("*.xlsx")):
        wb = load_workbook(path, read_only=False, data_only=False)
        for ws in wb.worksheets:
            if "ANALISE" not in normalize_label(ws.title):
                continue
            summary, daily, collaborators, issues = analyze_analysis_sheet(path, ws)
            summaries.append(summary)
            daily_rows.extend(daily)
            collaborator_rows.extend(collaborators)
            issues_rows.extend(
                {
                    "Arquivo": issue.arquivo,
                    "Aba": issue.aba,
                    "Celula": issue.celula,
                    "Problema": issue.problema,
                    "Encontrado": excel_text(issue.encontrado),
                    "Esperado": excel_text(issue.esperado),
                }
                for issue in issues
            )

    summaries_df = pd.DataFrame(summaries).sort_values("% real x teórica", ascending=False)
    daily_df = pd.DataFrame(daily_rows).sort_values(["Setor", "Data"])
    collaborators_df = pd.DataFrame(collaborator_rows).sort_values(["Setor", "Produção total"], ascending=[True, False])
    issues_df = pd.DataFrame(issues_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summaries_df.to_excel(writer, sheet_name="Consolidado", index=False)
        daily_df.to_excel(writer, sheet_name="Diario", index=False)
        collaborators_df.to_excel(writer, sheet_name="Colaboradores", index=False)
        issues_df.to_excel(writer, sheet_name="Inconsistencias", index=False)

    style_workbook(output_path)

    report_path.parent.mkdir(parents=True, exist_ok=True)
    write_report(report_path, summaries_df, issues_df)

    print(f"Resumo Excel: {output_path}")
    print(f"Relatório: {report_path}")


if __name__ == "__main__":
    main()
